#!python3
#将工资分配表里的发共享表格提取出来保存为单独的文件，并删除金额为0的行，将填充格式去掉

import openpyxl,os
from openpyxl.utils import get_column_letter,column_index_from_string#列数数字转字母，字母转数字

def replace_xlsx(wbname,exOriginal):#复制wbname的工作表到单独文件(工作表名称,工作薄名称)，并且删除金额为0的行
    os.chdir(exFilePath)    #更改工作目录
    wb=openpyxl.load_workbook(exOriginal,data_only=True)
    # 遍历wb的工作表名,删除名字不是wbname的
    for sheet in wb:
        if sheet.title!=wbname:
            wb.remove(sheet)
        
    ws=wb[wbname]
    isFindNumCol=False#判断是否取到了numcol值，即是否取到了金额列所在的列数
    isFindNumber=False#判断是否取得序号所在列数
    delRowList=[]
    #找到金额对应列,并且取得要删除行的行号列表    
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):#遍历row元组
        #找到金额对应列
        if isFindNumCol==False:#如果没找到值，进入循环开始找        
            for cellobj in row:
                if cellobj.value=='行项目编号\n（每张凭证从1开始顺序排列）\n必输' or cellobj.value=='行项目编号':
                    numberY=cellobj.column
                    numberX=cellobj.row
                    isFindNumber=True
                    print('工作表'+wbname+'编号列为'+str(numberY)+'，行为'+str(numberX))
                if cellobj.value=='凭证货币金额\n必输' or cellobj.value=='金额':
                    #print(cellobj.column,cellobj.value)
                    numcol=cellobj.column
                    # print('工作表'+wbname+'金额列为'+str(numcol))
                    isFindNumCol=True
                    print('工作表'+wbname+'金额列为'+str(numcol))
                    break
        #循环结束，如果在第一行有金额列，则获得了金额列所在的列数numcol
        
        # print(row[numcol-1].value)
        #判断是否获取到了值，没获取到值直接进入下一行的循环去找金额列
        if isFindNumCol==False and isFindNumber==False:
            continue
        elif isFindNumCol==True and isFindNumber==False:
            print('找到了金额列没找到序号列,检查表格或者代码吧')
            return
        elif isFindNumCol==False and isFindNumber==True:
            print('找到了序号列，没找到金额列，检查表格或者代码吧')
            return
        #row是个元组，包含了第row对应的行的cell对象。假设第一个循环，如果用列表的方式调用，row[0]代表A1的cell对象，row[2]代表B1的cell对象。调用值的话就用row[0].value
        if row[numcol-1].value==0 or row[numcol-1].value==None:#如果金额列对应的数值为0
            delRow=row[numcol-1].row#获取要删除金额的行号
            # print('删除'+str(delRow))
            delRowList.append(delRow)#获取要删除金额的列表
    
    #删除列表中的行
    delRowList.sort(reverse=True)#删除金额的列表重大到小排序，从后往前删除行，防止重前往后删除导致行号改变
    for iList in range(len(delRowList)):#遍历列表，删除对应的行
        # print(delRowList[iList])
        ws.delete_rows(delRowList[iList])
    #编号重新排序
    lineIndex=1
    for iNum in range(numberX,ws.max_row+1):
        # print(numberY)
        # print(ws[iNum][numberY])
        ws[iNum+1][numberY-1].value=lineIndex#通过ws[][]引用从0开始?
        lineIndex+=1

    
    
    # for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        # for cell in row:
            # print(cell.value)
    
    
    
    
    wb.save(wbname+'.xlsx')
    print('生成工作表'+wbname+'.xlsx')

# print(get_column_letter(ws.max_column))
# print(ws.max_row)
# maxrow=ws.max_row
# maxcolumn=ws.max_column

#判断金额列是否为0或者空，如果为零或者空就删除

#写入新的表格中
exMonth=input("保存你的工资表,并输入月份,如输入11则会在11月的文件夹中执行操作\n记得将共享-油建电气1911工资及附加自己分配表改.xlsx的19去掉\n")
exFilePath='E:\\财务资料\\`每月\\工资\\'+exMonth+'月'
exOriginal='共享-油建电气'+exMonth+'工资及附加自己分配表改.xlsx'
print('工作目录'+exFilePath+'\n文件名'+exOriginal)
wbNameList=['职工工资','公积金年金','工资三费','中友劳务','其他劳务','五项统筹']
for wbname in wbNameList:
    replace_xlsx(wbname,exOriginal)
# replace_xlsx('职工工资',exOriginal)